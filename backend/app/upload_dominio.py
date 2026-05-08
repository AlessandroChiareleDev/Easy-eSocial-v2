"""F8.1 — Upload XLSX Dominio (porte de V1 uploadController + database-service).

Endpoints:
    POST /api/dominio/upload            -> recebe XLSX, popula tabela_eb + analise_natureza
    GET  /api/dominio/uploads           -> lista uploads recentes
    GET  /api/dominio/tabela-eb         -> lista linhas (paginado)
    GET  /api/dominio/analise-natureza  -> lista linhas (paginado)

Multi-tenant via header `X-Empresa-CNPJ` (resolvido por tenant.empresa_conn).

V1 uploadava 4 abas: 'ANALISE NATUREZA', 'Dinamica', 'Tabela Eventos GI', 'Tabela EB'.
V2 mantem apenas as duas usadas (ANALISE NATUREZA -> analise_natureza, Tabela EB -> tabela_eb).
Casing/whitespace flexivel nos nomes de aba.
"""
from __future__ import annotations

import io
import json
import re
from typing import Any

from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook

from . import tenant


router = APIRouter(prefix="/api/dominio", tags=["dominio"])


# Limites
MAX_BYTES = 200 * 1024 * 1024            # 200 MB (mesmo do V1)
MAX_BATCH = 500                          # rows por INSERT VALUES (...)

# Schema de cada tabela landing.
# (n_cols, has_row_number, has_raw_data)
TABLE_SHAPE: dict[str, tuple[int, bool, bool]] = {
    "tabela_eb":         (26, False, True),   # col_a..col_z + raw_data
    "analise_natureza":  (10, True,  False),  # col_a..col_j + row_number
}


# ----------------------------------------------------------------- helpers

_NORM_RE = re.compile(r"\s+")


def _normalize_sheet_name(name: str) -> str:
    """Lowercase + colapsa whitespace + sem acentos basicos."""
    s = name.strip().lower()
    s = _NORM_RE.sub(" ", s)
    return (
        s.replace("á", "a").replace("é", "e").replace("í", "i")
         .replace("ó", "o").replace("ú", "u").replace("ç", "c")
         .replace("ã", "a").replace("õ", "o").replace("â", "a")
         .replace("ê", "e").replace("ô", "o")
    )


def _resolve_target_table(sheet_name: str) -> str | None:
    norm = _normalize_sheet_name(sheet_name)
    if norm in ("analise natureza", "analisenatureza"):
        return "analise_natureza"
    if norm in ("tabela eb", "tabelaeb"):
        return "tabela_eb"
    return None


def _col_name(idx: int) -> str:
    """0-based -> col_a, 25-> col_z, 26-> col_aa..."""
    letters = ""
    n = idx
    while n >= 0:
        letters = chr(ord("a") + (n % 26)) + letters
        n = n // 26 - 1
    return f"col_{letters}"


def _cnpj_ativo(request: Request) -> str:
    cnpj = getattr(request.state, "cnpj_ativo", None)
    if not cnpj:
        cnpj = request.headers.get("X-Empresa-CNPJ", "").strip() or None
    if not cnpj:
        raise HTTPException(400, "header X-Empresa-CNPJ ausente")
    return cnpj


def _cell_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        return v
    return str(v)


# ----------------------------------------------------------------- core

def _read_xlsx(content: bytes) -> dict[str, list[list[Any]]]:
    """Le todas as abas e devolve dict[name -> list-of-rows]. Preserva ordem original."""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f"XLSX invalido: {e}")
    sheets: dict[str, list[list[Any]]] = {}
    for ws in wb.worksheets:
        rows: list[list[Any]] = []
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
        sheets[ws.title] = rows
    return sheets


def _truncate_target(cur, table: str) -> None:
    cur.execute(f"TRUNCATE TABLE {table} RESTART IDENTITY CASCADE")


def _insert_rows(cur, table: str, rows: list[list[Any]]) -> tuple[int, int]:
    """Insere `rows` em `table` materializando col_a..col_<n_cols-1> + raw_data (+ row_number)."""
    if not rows:
        return 0, 0
    n_cols, has_row_num, has_raw = TABLE_SHAPE[table]
    col_names = [_col_name(i) for i in range(n_cols)]
    if has_raw:
        col_names.append("raw_data")
    if has_row_num:
        col_names.append("row_number")
    placeholder_row = "(" + ", ".join(["%s"] * len(col_names)) + ")"
    sql_prefix = f"INSERT INTO {table} ({', '.join(col_names)}) VALUES "

    inserted = 0
    for batch_start in range(0, len(rows), MAX_BATCH):
        batch = rows[batch_start:batch_start + MAX_BATCH]
        params: list[Any] = []
        placeholders: list[str] = []
        for row_idx, raw in enumerate(batch, start=batch_start + 1):
            cells = [_cell_str(raw[i]) if i < len(raw) else None for i in range(n_cols)]
            params.extend(cells)
            if has_raw:
                params.append(json.dumps([_cell_str(v) for v in raw], ensure_ascii=False))
            if has_row_num:
                params.append(row_idx)
            placeholders.append(placeholder_row)
        cur.execute(sql_prefix + ", ".join(placeholders), params)
        inserted += len(batch)
    return inserted, n_cols


# ----------------------------------------------------------------- endpoints

@router.post("/upload")
async def upload_dominio(request: Request, file: UploadFile = File(...)):
    cnpj = _cnpj_ativo(request)

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "envie .xlsx ou .xls")

    content = await file.read()
    if len(content) > MAX_BYTES:
        raise HTTPException(413, f"arquivo > {MAX_BYTES // 1024 // 1024}MB")
    if not content:
        raise HTTPException(400, "arquivo vazio")

    sheets = _read_xlsx(content)

    # Mapeia abas -> tabela alvo
    plan: list[tuple[str, str, list[list[Any]]]] = []  # (sheet_name, target_table, rows)
    for sheet_name, rows in sheets.items():
        target = _resolve_target_table(sheet_name)
        if target is None:
            continue
        # primeira linha = header — descarta na insercao mas preserva no count
        data_rows = rows[1:] if rows else []
        plan.append((sheet_name, target, data_rows))

    if not plan:
        raise HTTPException(
            422,
            f"nenhuma aba reconhecida. abas={list(sheets.keys())}. "
            f"esperado: 'ANALISE NATUREZA' e/ou 'Tabela EB'",
        )

    summary: dict[str, Any] = {
        "filename": file.filename,
        "size_bytes": len(content),
        "sheets_total": len(sheets),
        "sheets_ignoradas": [
            n for n in sheets if _resolve_target_table(n) is None
        ],
        "tabelas": [],
    }

    with tenant.empresa_conn(cnpj) as conn:
        try:
            with conn.cursor() as cur:
                # uma transacao atomica
                for sheet_name, target, data_rows in plan:
                    _truncate_target(cur, target)
                    inseridas, n_cols = _insert_rows(cur, target, data_rows)
                    summary["tabelas"].append({
                        "sheet_name": sheet_name,
                        "tabela": target,
                        "linhas_inseridas": inseridas,
                        "colunas_materializadas": n_cols,
                    })
            conn.commit()
        except HTTPException:
            conn.rollback()
            raise
        except Exception as e:  # noqa: BLE001
            conn.rollback()
            raise HTTPException(500, f"falha ao persistir: {e}")

    return {"ok": True, **summary}


@router.get("/tabela-eb")
def listar_tabela_eb(request: Request, limit: int = 100, offset: int = 0):
    cnpj = _cnpj_ativo(request)
    limit = max(1, min(limit, 1000))
    offset = max(0, offset)
    with tenant.empresa_conn(cnpj) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM tabela_eb")
            total = cur.fetchone()[0]
            cur.execute(
                "SELECT id, col_a, col_b, col_c, col_d, col_e, col_f, col_g, col_h, col_i, col_j "
                "FROM tabela_eb ORDER BY id LIMIT %s OFFSET %s",
                (limit, offset),
            )
            cols = [d.name for d in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    return {"total": total, "limit": limit, "offset": offset, "data": rows}


@router.get("/analise-natureza")
def listar_analise_natureza(
    request: Request,
    limit: int = 100,
    offset: int = 0,
    apenas_verificar: bool = False,
):
    cnpj = _cnpj_ativo(request)
    limit = max(1, min(limit, 1000))
    offset = max(0, offset)

    where = "WHERE UPPER(TRIM(col_d)) = 'VERIFICAR'" if apenas_verificar else ""
    with tenant.empresa_conn(cnpj) as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM analise_natureza {where}")
            total = cur.fetchone()[0]
            cur.execute(
                f"SELECT id, col_a, col_b, col_c, col_d, col_e, col_f "
                f"FROM analise_natureza {where} "
                f"ORDER BY id LIMIT %s OFFSET %s",
                (limit, offset),
            )
            cols = [d.name for d in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    return {"total": total, "limit": limit, "offset": offset, "data": rows}


@router.get("/resumo")
def resumo(request: Request):
    cnpj = _cnpj_ativo(request)
    with tenant.empresa_conn(cnpj) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM tabela_eb")
            tabela_eb = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM analise_natureza")
            analise = cur.fetchone()[0]
            cur.execute(
                "SELECT COUNT(*) FROM analise_natureza WHERE UPPER(TRIM(col_d))='VERIFICAR'"
            )
            verificar = cur.fetchone()[0]
    return {
        "cnpj": cnpj,
        "tabela_eb_rows": tabela_eb,
        "analise_natureza_rows": analise,
        "analise_natureza_verificar": verificar,
    }
