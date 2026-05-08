"""F8.4 — Cruzamento contabil (porte de V1 cruzamentoRoutes.ts).

Recebe XLSX com 2 abas, popula cruzamento_tabela_a/b, depois INNER JOIN por col_a -> cruzamento_resultado.
"""
from __future__ import annotations

import io
import json
from typing import Any

from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook

from . import tenant


router = APIRouter(prefix="/api/cruzamento", tags=["cruzamento"])

MAX_BYTES = 200 * 1024 * 1024
MAX_COLS = 52   # col_a..col_az
MAX_BATCH = 200


def _col_name(idx: int) -> str:
    letters = ""
    n = idx
    while n >= 0:
        letters = chr(ord("a") + (n % 26)) + letters
        n = n // 26 - 1
    return f"col_{letters}"


def _cell_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        return v
    return str(v)


def _cnpj_ativo(request: Request) -> str:
    cnpj = getattr(request.state, "cnpj_ativo", None)
    if not cnpj:
        cnpj = request.headers.get("X-Empresa-CNPJ", "").strip() or None
    if not cnpj:
        raise HTTPException(400, "header X-Empresa-CNPJ ausente")
    return cnpj


def _read_xlsx(content: bytes) -> list[tuple[str, list[list[Any]]]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f"XLSX invalido: {e}")
    sheets: list[tuple[str, list[list[Any]]]] = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        sheets.append((ws.title, rows))
    return sheets


def _is_blank_row(row: list[Any]) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def _insert_sheet(
    cur, table: str, upload_id: int, headers: list[str], data_rows: list[list[Any]]
) -> int:
    n_cols = min(MAX_COLS, len(headers))
    col_names = ["cruzamento_upload_id", "row_number"] + [_col_name(i) for i in range(n_cols)] + ["raw_data"]
    placeholder = "(" + ", ".join(["%s"] * len(col_names)) + ")"
    sql_prefix = f"INSERT INTO {table} ({', '.join(col_names)}) VALUES "

    inserted = 0
    pending: list[list[Any]] = []
    placeholders: list[str] = []

    def flush():
        nonlocal pending, placeholders, inserted
        if not pending:
            return
        flat: list[Any] = [v for row in pending for v in row]
        cur.execute(sql_prefix + ", ".join(placeholders), flat)
        inserted += len(pending)
        pending = []
        placeholders = []

    row_num = 0
    for raw in data_rows:
        if _is_blank_row(raw):
            continue
        row_num += 1
        cells = [_cell_str(raw[i]) if i < len(raw) else None for i in range(n_cols)]
        raw_obj = {}
        for c, val in enumerate(raw):
            key = headers[c] if c < len(headers) and headers[c] else f"Col{c}"
            raw_obj[key] = _cell_str(val)
        params = [upload_id, row_num] + cells + [json.dumps(raw_obj, ensure_ascii=False)]
        pending.append(params)
        placeholders.append(placeholder)
        if len(pending) >= MAX_BATCH:
            flush()
    flush()
    return inserted


@router.post("/upload")
async def upload_cruzamento(request: Request, file: UploadFile = File(...)):
    cnpj = _cnpj_ativo(request)
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "envie .xlsx ou .xls")
    content = await file.read()
    if len(content) > MAX_BYTES:
        raise HTTPException(413, f"arquivo > {MAX_BYTES // 1024 // 1024}MB")
    if not content:
        raise HTTPException(400, "arquivo vazio")

    sheets = _read_xlsx(content)
    if len(sheets) < 2:
        raise HTTPException(400, f"precisa de >= 2 abas. encontradas={len(sheets)}: {[s[0] for s in sheets]}")

    sheet_names = [s[0] for s in sheets]

    with tenant.empresa_conn(cnpj) as conn:
        try:
            with conn.cursor() as cur:
                # limpar
                cur.execute("DELETE FROM cruzamento_resultado")
                cur.execute("DELETE FROM cruzamento_tabela_b")
                cur.execute("DELETE FROM cruzamento_tabela_a")
                cur.execute("DELETE FROM cruzamento_uploads")

                cur.execute(
                    "INSERT INTO cruzamento_uploads (filename, original_name, file_size, sheet_count, sheet_names) "
                    "VALUES (%s, %s, %s, %s, %s) RETURNING id",
                    (file.filename, file.filename, len(content), len(sheets), json.dumps(sheet_names)),
                )
                upload_id = cur.fetchone()[0]

                tables = ["cruzamento_tabela_a", "cruzamento_tabela_b"]
                resumo = []
                for idx in range(2):
                    name, rows = sheets[idx]
                    headers = [str(h) if h is not None else "" for h in (rows[0] if rows else [])]
                    data = rows[1:] if rows else []
                    inseridas = _insert_sheet(cur, tables[idx], upload_id, headers, data)
                    resumo.append({
                        "sheet_name": name,
                        "tabela": tables[idx],
                        "linhas_inseridas": inseridas,
                        "headers": headers,
                    })
            conn.commit()
        except HTTPException:
            conn.rollback()
            raise
        except Exception as e:  # noqa: BLE001
            conn.rollback()
            raise HTTPException(500, f"falha: {e}")

    return {"ok": True, "upload_id": upload_id, "sheets": resumo}


@router.post("/cruzar")
def cruzar(request: Request):
    """INNER JOIN A x B por TRIM(col_a). Extrai INSS/IRRF/FGTS de B.raw_data por header."""
    cnpj = _cnpj_ativo(request)
    with tenant.empresa_conn(cnpj) as conn:
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM cruzamento_uploads ORDER BY id DESC LIMIT 1")
                row = cur.fetchone()
                if not row:
                    raise HTTPException(400, "Nenhum upload encontrado. envie o arquivo primeiro.")
                upload_id = row[0]

                cur.execute("SELECT raw_data FROM cruzamento_tabela_b LIMIT 1")
                sample = cur.fetchone()
                if not sample:
                    raise HTTPException(400, "Tabela B vazia.")
                raw_sample = sample[0] if isinstance(sample[0], dict) else json.loads(sample[0])
                keys = list(raw_sample.keys())

                def find_key(pattern: str) -> str:
                    p = pattern.lower()
                    for k in keys:
                        if p in k.lower():
                            return k
                    return ""

                inss_key = find_key("inss")
                irrf_key = find_key("irrf")
                fgts_key = find_key("fgts")

                cur.execute("DELETE FROM cruzamento_resultado")
                cur.execute(
                    """
                    INSERT INTO cruzamento_resultado
                        (cruzamento_upload_id, codigo, nome_evento, natureza_esocial,
                         cod_inss, cod_irrf, cod_fgts, row_number)
                    SELECT
                        %s,
                        TRIM(a.col_a),
                        a.col_b,
                        a.col_c,
                        b.raw_data->>%s,
                        b.raw_data->>%s,
                        b.raw_data->>%s,
                        ROW_NUMBER() OVER (ORDER BY
                            CASE WHEN TRIM(a.col_a) ~ '^[0-9]+$'
                                 THEN CAST(TRIM(a.col_a) AS INTEGER)
                                 ELSE NULL END NULLS LAST,
                            TRIM(a.col_a))
                    FROM cruzamento_tabela_a a
                    INNER JOIN cruzamento_tabela_b b ON TRIM(a.col_a) = TRIM(b.col_a)
                    """,
                    (upload_id, inss_key, irrf_key, fgts_key),
                )
                total = cur.rowcount
            conn.commit()
        except HTTPException:
            conn.rollback()
            raise
        except Exception as e:  # noqa: BLE001
            conn.rollback()
            raise HTTPException(500, f"falha: {e}")
    return {
        "ok": True,
        "total": total,
        "keys_detectadas": {"inss": inss_key, "irrf": irrf_key, "fgts": fgts_key},
    }


@router.get("/status")
def status(request: Request):
    cnpj = _cnpj_ativo(request)
    with tenant.empresa_conn(cnpj) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT cu.id, cu.original_name, cu.file_size, cu.sheet_count, cu.sheet_names, cu.criado_em,
                       (SELECT COUNT(*) FROM cruzamento_tabela_a WHERE cruzamento_upload_id = cu.id),
                       (SELECT COUNT(*) FROM cruzamento_tabela_b WHERE cruzamento_upload_id = cu.id),
                       (SELECT COUNT(*) FROM cruzamento_resultado WHERE cruzamento_upload_id = cu.id)
                  FROM cruzamento_uploads cu
                 ORDER BY cu.id DESC LIMIT 1
                """
            )
            row = cur.fetchone()
    if not row:
        return {"ok": True, "has_data": False}
    return {
        "ok": True,
        "has_data": True,
        "upload": {
            "id": row[0],
            "original_name": row[1],
            "file_size": row[2],
            "sheet_count": row[3],
            "sheet_names": row[4] if isinstance(row[4], list) else json.loads(row[4]) if row[4] else [],
            "criado_em": row[5].isoformat() if row[5] else None,
            "rows_a": row[6],
            "rows_b": row[7],
            "rows_resultado": row[8],
        },
    }


@router.get("/resultado")
def listar_resultado(request: Request, limit: int = 100, offset: int = 0):
    cnpj = _cnpj_ativo(request)
    limit = max(1, min(limit, 1000))
    offset = max(0, offset)
    with tenant.empresa_conn(cnpj) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM cruzamento_resultado")
            total = cur.fetchone()[0]
            cur.execute(
                "SELECT codigo, nome_evento, natureza_esocial, cod_inss, cod_irrf, cod_fgts "
                "FROM cruzamento_resultado ORDER BY row_number LIMIT %s OFFSET %s",
                (limit, offset),
            )
            cols = [d.name for d in cur.description]
            data = [dict(zip(cols, r)) for r in cur.fetchall()]
    return {"total": total, "limit": limit, "offset": offset, "data": data}
