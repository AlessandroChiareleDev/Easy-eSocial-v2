from __future__ import annotations

import argparse
import re
import shutil
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import psycopg2.extras
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))

from app import db, storage, tenant  # noqa: E402


EMPRESA_ID = tenant.SOLUCOES_ID
STATUS_EXPORTADO = "sucesso"
MESES_PT = {
    "01": "janeiro",
    "02": "fevereiro",
    "03": "marco",
    "04": "abril",
    "05": "maio",
    "06": "junho",
    "07": "julho",
    "08": "agosto",
    "09": "setembro",
    "10": "outubro",
    "11": "novembro",
    "12": "dezembro",
}


def safe_name(value: str) -> str:
    value = (value or "").strip()
    value = re.sub(r"[^A-Za-z0-9_.-]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("._")
    return value or "sem_nome"


def format_cpf(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    if len(digits) != 11:
        return value or ""
    return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"


def event_id_and_name(xml_bytes: bytes, fallback: str) -> tuple[str, str]:
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return fallback, ""

    event_id = ""
    nome = ""
    for elem in root.iter():
        if not event_id:
            candidate = elem.attrib.get("Id") or elem.attrib.get("id")
            if candidate:
                event_id = candidate.strip()
        local = elem.tag.rsplit("}", 1)[-1]
        if local in {"nmTrab", "nmBenef", "nome"} and (elem.text or "").strip():
            nome = (elem.text or "").strip()
        if event_id and nome:
            break
    return event_id or fallback, nome


def read_lo(conn, oid: int) -> bytes:
    return b"".join(storage.iter_lo_bytes(conn, int(oid)))


def rows_for_month(conn, per_apur: str) -> list[dict]:
    internal_id = tenant.internal_empresa_id(EMPRESA_ID)
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(
            """
            SELECT it.id AS item_id,
                   it.timeline_envio_id,
                   it.cpf,
                   it.status,
                   it.nr_recibo_novo,
                   it.xml_enviado_oid,
                   it.criado_em,
                   te.sequencia,
                    tm.per_apur,
                    lom.oid IS NOT NULL AS xml_enviado_lo_existe
              FROM timeline_envio_item it
              JOIN timeline_envio te ON te.id = it.timeline_envio_id
              JOIN timeline_mes tm ON tm.id = te.timeline_mes_id
                LEFT JOIN pg_catalog.pg_largeobject_metadata lom ON lom.oid = it.xml_enviado_oid
             WHERE tm.empresa_id = %s
               AND tm.per_apur = %s
               AND it.tipo_evento = 'S-1210'
               AND it.status = %s
               AND it.xml_enviado_oid IS NOT NULL
             ORDER BY it.cpf, it.criado_em, it.id
            """,
            (internal_id, per_apur, STATUS_EXPORTADO),
        )
        return [dict(r) for r in cur.fetchall()]


def write_xlsx(path: Path, mapping_rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "depara"
    headers = [
        "cpf",
        "cpf_normalizado",
        "nome_xml",
        "recibo_novo",
        "timeline_envio_id",
        "item_id",
    ]
    ws.append(headers)
    for row in mapping_rows:
        ws.append([
            format_cpf(row.get("cpf") or ""),
            row.get("cpf"),
            row.get("nome_xml"),
            row.get("nr_recibo_novo"),
            row.get("timeline_envio_id"),
            row.get("item_id"),
        ])
    ws.freeze_panes = "A2"
    widths = [16, 18, 72, 30, 18, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def write_missing_xlsx(path: Path, missing_rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "xmls_ausentes"
    headers = [
        "cpf",
        "cpf_normalizado",
        "recibo_novo",
        "timeline_envio_id",
        "item_id",
        "xml_enviado_oid",
    ]
    ws.append(headers)
    for row in missing_rows:
        ws.append([
            format_cpf(row.get("cpf") or ""),
            row.get("cpf"),
            row.get("nr_recibo_novo"),
            row.get("timeline_envio_id"),
            row.get("item_id"),
            row.get("xml_enviado_oid"),
        ])
    ws.freeze_panes = "A2"
    widths = [16, 18, 30, 18, 12, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def zip_month(month_dir: Path, archive_base: Path, zip_path: Path) -> None:
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for path in sorted(month_dir.rglob("*")):
            if path.is_file() and not path.name.startswith(".~lock."):
                zf.write(path, path.relative_to(archive_base))


def export_month(per_apur: str, output_base: Path, overwrite: bool, allow_missing_lo: bool) -> dict:
    mes = per_apur[-2:]
    mes_nome = MESES_PT.get(mes, mes)
    root_dir = output_base / "relatório XML final soluções"
    month_dir = root_dir / f"{mes} - {mes_nome}"

    if month_dir.exists() and overwrite:
        shutil.rmtree(month_dir)
    month_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = month_dir / f"depara_s1210_enviados_{per_apur}.xlsx"
    missing_xlsx_path = month_dir / f"xmls_ausentes_{per_apur}.xlsx"
    zip_path = output_base / f"relatório XML final soluções - {per_apur}.zip"

    conn = db.connect(empresa_id=EMPRESA_ID)
    mapping_rows: list[dict] = []
    missing_rows: list[dict] = []
    used_names: set[str] = set()
    try:
        rows = rows_for_month(conn, per_apur)
        total = len(rows)
        missing_total = sum(1 for row in rows if not row.get("xml_enviado_lo_existe"))
        if missing_total and not allow_missing_lo:
            raise RuntimeError(
                f"{per_apur}: {missing_total}/{total} XMLs apontam para Large Objects ausentes. "
                "Use --allow-missing-lo para exportar somente os XMLs preservados e gerar auditoria."
            )
        print(f"[INFO] {per_apur}: {total} XMLs S-1210 com status={STATUS_EXPORTADO}")
        if missing_total:
            print(f"[WARN] {per_apur}: {missing_total} XMLs com Large Object ausente")
        for idx, row in enumerate(rows, start=1):
            if not row.get("xml_enviado_lo_existe"):
                mapping_rows.append({**row, "nome_xml": ""})
                missing_rows.append(row)
                if idx % 1000 == 0 or idx == total:
                    print(f"[INFO] processados {idx}/{total}")
                continue
            fallback_id = f"S1210_{per_apur}_item_{row['item_id']}"
            xml_bytes = read_lo(conn, int(row["xml_enviado_oid"]))
            event_id, nome = event_id_and_name(xml_bytes, fallback_id)
            base_name = safe_name(event_id)
            filename = f"{base_name}.xml"
            if filename in used_names:
                filename = f"{base_name}_item_{row['item_id']}.xml"
            used_names.add(filename)

            (month_dir / filename).write_bytes(xml_bytes)
            mapping_rows.append({**row, "nome_xml": filename})
            if idx % 1000 == 0 or idx == total:
                print(f"[INFO] processados {idx}/{total}")
        conn.rollback()
    finally:
        conn.close()

    write_xlsx(xlsx_path, mapping_rows)
    if missing_rows:
        write_missing_xlsx(missing_xlsx_path, missing_rows)
    zip_month(month_dir, output_base, zip_path)
    return {
        "root_dir": str(root_dir),
        "month_dir": str(month_dir),
        "xlsx": str(xlsx_path),
        "missing_xlsx": str(missing_xlsx_path) if missing_rows else "",
        "zip": str(zip_path),
        "xmls": len(mapping_rows) - len(missing_rows),
        "rows": len(mapping_rows),
        "missing_xmls": len(missing_rows),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--per-apur", default="2025-02")
    parser.add_argument("--output", default=str(ROOT / "relatorios"))
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--allow-missing-lo", action="store_true")
    args = parser.parse_args()

    output = Path(args.output).resolve()
    output.mkdir(parents=True, exist_ok=True)
    started = datetime.now()
    result = export_month(args.per_apur, output, args.overwrite, args.allow_missing_lo)
    elapsed = (datetime.now() - started).total_seconds()
    print("[OK] export concluido")
    print(f"[OK] xmls={result['xmls']} segundos={elapsed:.1f}")
    print(f"[OK] linhas_xlsx={result['rows']}")
    if result["missing_xmls"]:
        print(f"[WARN] xmls_ausentes={result['missing_xmls']}")
    print(f"[OK] pasta={result['month_dir']}")
    print(f"[OK] xlsx={result['xlsx']}")
    if result["missing_xlsx"]:
        print(f"[OK] auditoria_ausentes={result['missing_xlsx']}")
    print(f"[OK] zip={result['zip']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())